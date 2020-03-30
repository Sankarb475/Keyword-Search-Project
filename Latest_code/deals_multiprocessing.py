# -*- coding: utf-8 -*-
"""
Created on Mon Jan 20 12:36:08 2020
@author: sbiswas149
"""
import comtypes.client as coms
import win32com.client
import csv
import xlrd
import re
import xlsxwriter
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from pyxlsb import open_workbook
from independentsoft.msg import Message
from striprtf.striprtf import rtf_to_text
from wand.image import Image
from odf import text
from odf.opendocument import load
from sqlalchemy import create_engine
import ast
import configparser
import textract
from pdf2image import convert_from_path, convert_from_bytes
import logging
import logging.config
import os, sys
from PIL import Image
from visio2pdf import Visio2PDFConverter
import datetime
import multiprocessing

# create logger
logger = logging.getLogger('VDR')
logging.basicConfig(filename= r'D:\Backend Python\Scripts\logs\example.log',
                            filemode='a',
                            format='%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s',
                            datefmt='%H:%M:%S',
                            level=logging.DEBUG)
logging.warning('is when this event was logged.')
outputCol = 0
outputRow = 1
memory_error_files = []
exception_files = []
keyword_dic = {}

def keyword_extraction():
    try:
        db_connection_str = 'mysql+pymysql://root:admin@localhost/virtual_data_room'
        db_connection = create_engine(db_connection_str)
        df = pd.read_sql('SELECT keyword, category FROM data_dictionary', con=db_connection)
        return df.drop_duplicates()
    except:
        print("mysql connection issue")
        sys.exit(1)
df = keyword_extraction()
category_list = df["category"].drop_duplicates().tolist()
for i in category_list:
    temp = df[df["category"] == i]["keyword"].tolist()
    keyword_dic[i] = temp
pattern_list = []
try:
    config = configparser.ConfigParser()
    config.read('properties.ini')
    output_file = config['Directories']['output_file']
    temp_storage_path = config['Directories']['temp_storage']
    keyword_file = config['Directories']['keyword_file']
    intermediate_file = config['Directories']['intermediate_file']
except KeyError as e:
    print("Config file is empty")
    logger.critical('Config file is empty - fatal error')
    sys.exit(1)
except Exception as e:
    print("Check config file", e)
    sys.exit(1)

workbook = xlsxwriter.Workbook(output_file)
# sheet - Raw Details
header_format = workbook.add_format({
    'bold': True,
    'text_wrap': True,
    'valign': 'top',
    'fg_color': '#FFFF00',
    'border': 1})

worksheet1 = workbook.add_worksheet("Raw Details")
worksheet1.set_column(0, 8, 25)
worksheet1.write('A1', 'File Name', header_format)
worksheet1.write('B1', 'Index Number', header_format)
worksheet1.write('C1', 'Folder Name', header_format)
worksheet1.write('D1', 'File Path', header_format)
worksheet1.write('E1', 'File Type', header_format)
worksheet1.write('F1', 'Total Number of Page(s)/Slide(s)', header_format)
worksheet1.write('G1', 'Keyword', header_format)
worksheet1.write('H1', 'IT Category', header_format)
worksheet1.write('I1', 'Page/Slide Number', header_format)


def specialCharReplace(word):
    return re.sub('[^A-Za-z0-9\s]+', '', word)

def cell_header(a):
    if a <= 26:
        return chr(a + 64)
    count = a // 26 + 64
    rest = a % 26 + 64
    return chr(count) + chr(rest)


def csv_handler(path):
    with open(path) as csvfile:
        readCSV = csv.reader(csvfile)
        count = 0
        page_count = 1
        for row in readCSV:
            # count is row number
            count = count + 1
            for i in range(len(row)):
                # print(row[i] + " " + str(count) + " " + str(i))
                # row is representing each row  of the csv file - a list
                list_details = folderName(path)
                cell_details = cell_header(i + 1) + str(count)
                sheet_handler(list_details, cell_details, path, page_count, row[i])


def pattern(keys):
    return re.compile('|'.join([r'(?<!\w)%s(?!\w)' % re.escape(keys)]), flags=re.I)

def sheet_handler(list_details, cell_details, path, page_count, row):
    try:
        if isinstance(row, str):
            for category, keywords in keyword_dic.items():
                for keys in keywords:
                    print(row)
                    r = pattern(keys)
                    words = r.findall(row)
                    print("words", words)
                    for i in words:
                        print("Here for the excels")
                        outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3],
                                              page_count, keys, category, cell_details)

    except Exception as e:
        print("Error while writing to RAW details", e)


def removal(text):
    return text.replace("\\r", " ").replace("\\n", " ").replace('\\x', " ").replace('0c', " ")

def img_handler(path):
    list_details = folderName(path)
    text = textract.process(path, encoding='ascii', method='tesseract')
    content = removal(str(text))
    print(content)
    sheet_handler(list_details, 1, path, 1, content)

def visio_handler(path):
    out_file = os.path.join(temp_storage_path, "temp_visio.pdf")
    vpc = Visio2PDFConverter(visio_process_name='visio.exe',
                             current_working_directory=None,
                             temp_file_name='visio2pdf4latex_temp', visio_ext_names=['vsdx', 'vsd'])
    vpc.convert('one_visio_file.vsd')

def tiff_handler(path):
    list_details = folderName(path)
    out_file = os.path.join(temp_storage_path, "temp_img.JPEG")
    im = Image.open(path)
    out = im.convert("RGB")
    out.save(out_file, "JPEG", quality=90)
    img_handler_for_pdf(out_file, path, 1, 1, list_details)
    os.remove(out_file)

def doc_handler(path):
    #word = comtypes.client.CreateObject('Word.Application')
    wdFormatPDF = 17
    out_file = os.path.join(temp_storage_path, "temp_doc.pdf")
    word = win32com.client.DispatchEx("Word.Application")
    doc = word.Documents.Open(path)
    doc.SaveAs(out_file, FileFormat=wdFormatPDF)
    doc.Close()
    word.Quit()
    list_details = folderName(path)
    pdf_handler_for_doc(out_file, path, list_details)
    os.remove(out_file)

def ppt_handler(path, formatType = 32):
    powerpoint = win32com.client.DispatchEx("Powerpoint.Application")
    powerpoint.Visible = 1
    out_file = os.path.join(temp_storage_path, "temp_ppt.pdf")
    deck = powerpoint.Presentations.Open(path)
    deck.SaveAs(out_file, formatType) # formatType = 32 for ppt to pdf
    deck.Close()
    powerpoint.Quit()
    list_details = folderName(path)
    pdf_handler_for_doc(out_file, path, list_details)
    os.remove(out_file)


def pdf_handler_for_doc(pdf_path, path, list_details):
    #convertedpdf = pdf2image.convert_from_path(file,dpi=200,grayscale=False,poppler_path="C:/bin/",output_folder=savepath,fmt='jpg')
    #pages = convert_from_path(pdf_path, 500)
    try:
        source = Image(filename=os.path.normpath(pdf_path), resolution=120)
        total_pages = len(source.sequence)
        for index, image in enumerate(source.sequence):
            temp_file = list_details[0] + str(index + 1) + '.jpeg'
            temp_path = os.path.join(temp_storage_path, temp_file)
            Image(image).save(filename=temp_path)
            img_handler_for_pdf(temp_path, path, index, total_pages, list_details)
            os.remove(temp_path)
    except Exception as e:
        print("error", e)
        pass


def pdf_handler(path):
    try:
        list_details = folderName(path)
        source = Image(filename=path, resolution=120)
        total_pages = len(source.sequence)
        for index, image in enumerate(source.sequence):
            temp_file = list_details[0] + str(index + 1) + '.jpeg'
            temp_path = os.path.join(temp_storage_path, temp_file)
            Image(image).save(filename=temp_path)
            img_handler_for_pdf(temp_path, path, index, total_pages, list_details)
            os.remove(temp_path)

    except Exception as e:
        print("error", e)
        pass

def img_handler_for_pdf(temp_path, path, page_number ,total_pages, list_details):
    try:
        print("I am here pic")
        text = textract.process(temp_path, encoding="ascii",  errors='ignore', method='tesseract')
        content = removal(str(text))
        sheet_handler(list_details, page_number, path, total_pages, content)
    except MemoryError as e:
        memory_error_files.append(path)
        pass
    except Exception as e:
        exception_files.append(path)
        pass

def xlsx_handler(path):
    print("here inside xlsx")
    sheets_dict = pd.read_excel(path, sheet_name=None, header=None)
    sheet_count = len(sheets_dict)
    list_details = folderName(path)
    for name, sheet in sheets_dict.items():
        for index, row in sheet.iterrows():
            for i in range(len(row)):
                cell = cell_header(i + 1) + str(index + 1)
                cell_details = name + ": " + cell
                sheet_handler(list_details, cell_details, path, sheet_count, row[i])

def xlsb_handler(path):
    sheet_count = len(open_workbook(path).sheets)
    list_details = folderName(path)
    with open_workbook(path) as wb:
        for sheetname in wb.sheets:
            row_number = 0
            with wb.get_sheet(sheetname) as sheet:
                for row in sheet.rows():
                    row_number = row_number + 1
                    for i in range(len(row)):
                        if row[i].v:
                            cell = cell_header(i + 1) + str(row_number)
                            cell_details = sheetname + ": " + cell
                            sheet_handler(list_details, cell_details, path, sheet_count, row[i].v)


def odt_handler(path):
    textdoc = load(path)
    allparas = textdoc.getElementsByType(text.P)
    list_details = folderName(path)
    for i in allparas:
        sheet_handler(list_details, 1, path, 1, str(i))

def msg_handler(path):
    content = str(Message(path).body)
    sub = str(Message(path).subject)
    list_details = folderName(path)
    sheet_handler(list_details, 1, path, 1, content)
    sheet_handler(list_details, 1, path, 1, sub)


def txt_handler(path):
    list_details = folderName(path)
    file = open(path, mode='r')
    content = file.read()
    sheet_handler(list_details, 1, path, 1, content)
    file.close()


def rtf_handler(path):
    list_details = folderName(path)
    file = open(path, mode='r')
    rtf = file.read()
    content = rtf_to_text(rtf)
    sheet_handler(list_details, 1, path, 1, content)
    file.close()


def index_number_verification(index):
    r = re.compile('[@_!#$%^&*()<>?/\|}{~:]')
    if (index.isupper() or index.islower()) or r.search(index) != None:
        return " "
    if "_" in index:
        index = index.split("_")[0]
    index = re.sub('[^0-9\.\s]+', '', index)
    if len(index) == 1:
        index = re.sub('[.]', " ", index)
    return index


def folderName(path):
    # C:\Users\sb512911\Desktop\All\Screenshots\1.2.1 ILS Structure Chart.csv
    file_name = path.split("\\")[-1]
    index = file_name.split(" ")[0]
    index_number = index_number_verification(index)
    folder_name = path.split(file_name)[0]
    file_type = path.split(".")[-1]
    return [file_name, index_number, folder_name, file_type]


def outputWriterRawDetail(file_name, index_number, folder_name, file_path, file_type, pages, keyword, category, cell):
    global outputRow
    print("I am here")
    worksheet1.write(outputRow, 0, file_name)
    worksheet1.write(outputRow, 1, index_number)
    worksheet1.write(outputRow, 2, folder_name)
    worksheet1.write(outputRow, 3, file_path)
    worksheet1.write(outputRow, 4, file_type)
    worksheet1.write(outputRow, 5, pages)
    worksheet1.write(outputRow, 6, keyword)
    worksheet1.write(outputRow, 7, category)
    worksheet1.write(outputRow, 8, cell)
    outputRow = outputRow + 1


# pass the additional parameter for relevance test - dynamically
def generatingFileLevelDetail(file_path):
    try:
        data = pd.read_excel(open(file_path, 'rb'), sheet_name=0)
        if not data.empty:
            temp_data = data[["File Name", "Index Number", "Folder Name", "File Type"]]
            temp_data["Scan Status"] = "Successful"
            file_level_detail = temp_data[temp_data['File Type'] != "Unknown"] \
                .drop_duplicates().sort_values('File Name')

            file_level_detail["Number of Keywords found"] = \
                data[data["File Type"] != "Unknown"][['File Name']].sort_values('File Name').groupby(
                    ['File Name']).size().reset_index()[0].values

            file_level_detail["Number of Keyword Categories"] = \
                data[data["File Type"] != "Unknown"][['File Name', 'IT Category']].drop_duplicates().groupby(
                    ['File Name']).size() \
                    .reset_index()[0].values

            intermediate1 = data[data["File Type"] != "Unknown"][['File Name', 'Keyword']].groupby(
                ['File Name', 'Keyword']).size().reset_index()
            intermediate1[[0]] = intermediate1[[0]].astype(str).values
            intermediate1['joined'] = intermediate1[['Keyword', 0]].apply(lambda x: '('.join(x) + "), ", axis=1).values
            file_level_detail["Keyword Details"] = intermediate1[["File Name", "joined"]].groupby(
                ["File Name"]).sum().values

            file_level_detail["Category Details"] = \
                data[data["File Type"] != "Unknown"].groupby("File Name").apply(
                    lambda x: x["IT Category"].drop_duplicates().str.cat(sep=", ")).values

            file_level_detail["File Path"] = data[data["File Type"] != "Unknown"]["File Path"].drop_duplicates().values

            # Relevance check -- code here - assuming relevant count = 5
            mapping = {True: "Yes", False: "No"}
            file_level_detail["Relevant?"] = \
            data[data["File Type"] != "Unknown"][["File Name", "Keyword"]].drop_duplicates()[["File Name"]].groupby(
                ["File Name"]).size(). \
                reset_index()[0].apply(lambda x: x > 5).map(mapping).values

            temp_data_unknown = temp_data[["File Name", "Index Number", "Folder Name", "File Type"]][
                temp_data['File Type'] == "Unknown"]
            temp_dict = {"Number of Keywords found": 0, "Number of Keyword Categories": 0, "Keyword Details": "", \
                         "Category Details": "", "Relevant?": "", "File Path": "", "Scan Status": "Unsuccessful"}
            temp_df = pd.DataFrame([temp_dict])
            unknown_files = cross_join(temp_data_unknown, temp_df)
            file_level_detail = file_level_detail.append(unknown_files, ignore_index=True)
            writer = pd.ExcelWriter(output_file)
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                file_level_detail.to_excel(writer, "File Level Detail", index=False)
            writer.save()

            wb1 = openpyxl.load_workbook(output_file)
            ws = wb1['File Level Detail']
            fillBack = PatternFill(start_color="00FFFF00", fill_type="solid")

            ws.column_dimensions['A'].width = 24.94
            ws.column_dimensions['B'].width = 24.94
            ws.column_dimensions['C'].width = 24.94
            ws.column_dimensions['D'].width = 24.94
            ws.column_dimensions['E'].width = 24.94
            ws.column_dimensions['F'].width = 24.94
            ws.column_dimensions['G'].width = 24.94
            ws.column_dimensions['H'].width = 24.94
            ws.column_dimensions['I'].width = 24.94
            ws.column_dimensions['J'].width = 24.94
            ws.column_dimensions['K'].width = 24.94
            ws.row_dimensions[1].height = 28.8

            for cell in ws["1:1"]:
                cell.fill = fillBack
            wb1.save(output_file)

    except:
        print("error occured while generating file level detail")
        sys.exit(1)


def generatingKeywordDetails(file_path):
    data = pd.read_excel(open(file_path, 'rb'), sheet_name="Raw Details")
    try:
        if not data.empty:
            # distinct keywords and categories
            keyword_details = data[["Keyword", "IT Category"]].drop_duplicates()

            #  Number of files with keywords
            keyword_details = data[["Keyword", "File Name"]].drop_duplicates().groupby(["Keyword"]) \
                .size().reset_index().merge(keyword_details, on="Keyword")

            # renaming
            keyword_details = keyword_details. \
                rename(columns={'IT Category': 'Keyword Category', 0: 'Number of files with keywords'})

            # Total keyword hits
            keyword_details = data[["File Name", "Keyword"]].groupby(["Keyword"]) \
                .size().reset_index().merge(keyword_details, on="Keyword").rename(columns={0: "Total keyword hits"})

            # File(s) with maximum keyword hits and File Path(s)
            intermediate2 = data[["Keyword", "File Name", "File Path"]].groupby(["Keyword", "File Name", "File Path"]) \
                .size().reset_index().merge(keyword_details, on="Keyword").rename({0: "Count"}, axis=1) \
                [["File Name", "Keyword", "File Path", "Count"]]

            intermediate3 = intermediate2.groupby('Keyword')['Count'].apply(lambda x: x.eq(x.max()))

            intermediate4 = intermediate2.loc[intermediate3].groupby(['Keyword'])['File Name'].agg(
                ', '.join).reset_index()

            intermediate5 = intermediate2.loc[intermediate3].groupby(['Keyword'])['File Path'].agg(
                ', '.join).reset_index()

            keyword_details = keyword_details.merge(intermediate4, on="Keyword")

            keyword_details = keyword_details.merge(intermediate5, on="Keyword")

            keyword_details = keyword_details.rename(columns={"File Path": "File Path(s) with maximum keyword", \
                                                              "File(s) with maximum keyword hits": "File Path(s) with maximum keyword"})

            # All file(s) with keyword hits
            keyword_details["File Path(s)"] = \
                data[["Keyword", "File Path"]].groupby(["Keyword"]) \
                    .apply(lambda x: x["File Path"].drop_duplicates().str.cat(sep=", ")).reset_index() \
                    .merge(keyword_details, on="Keyword")[[0]].values

            # Index Number(s)
            keyword_details["Index Number(s)"] = \
                data[["Keyword", "Index Number"]].groupby("Keyword") \
                    .apply(lambda x: x["Index Number"].drop_duplicates().astype(str).str.cat(sep=", ")).reset_index() \
                    .merge(keyword_details, on="Keyword")[[0]].values

            # All the file names which has the keyword
            keyword_details["File Name(s)"] = \
                data[["Keyword", "File Name"]].groupby("Keyword") \
                    .apply(lambda x: x["File Name"].drop_duplicates().str.cat(sep=", ")).reset_index() \
                    .merge(keyword_details, on="Keyword")[[0]].values

            keyword_details = keyword_details.rename(columns={"File Name": "File Name(s) with maximum keyword", \
                                                              "File Name(s)": "All file(s) with keyword hits"})

            keyword_details = keyword_details[["Keyword", "Keyword Category", "Number of files with keywords", \
                                               "Total keyword hits", "File Name(s) with maximum keyword",
                                               "File Path(s) with maximum keyword", \
                                               "All file(s) with keyword hits", "Index Number(s)", "File Path(s)"]]

            # writing to "File Level Detail" sheet
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                keyword_details.to_excel(writer, "Keyword Details", index=False)
            writer.save()

            wb1 = openpyxl.load_workbook(output_file)
            ws = wb1['Keyword Details']
            fillBack = PatternFill(start_color="00FFFF00", fill_type="solid")

            ws.column_dimensions['A'].width = 24.94
            ws.column_dimensions['B'].width = 24.94
            ws.column_dimensions['C'].width = 24.94
            ws.column_dimensions['D'].width = 24.94
            ws.column_dimensions['E'].width = 24.94
            ws.column_dimensions['F'].width = 24.94
            ws.column_dimensions['G'].width = 24.94
            ws.column_dimensions['H'].width = 24.94
            ws.column_dimensions['I'].width = 24.94
            ws.row_dimensions[1].height = 28.8

            for cell in ws["1:1"]:
                cell.fill = fillBack
            wb1.save(output_file)

    except Exception as e:
        print("Error occured while generating Keyword details", e)
        sys.exit(1)


def generatingDRLDetails(drl_path, output_path):
    df = pd.read_excel(output_path, sheet_name="Raw Details")
    data = df[df["File Type"] != "Unknown"][["File Name", "Index Number", "File Path", "Keyword"]].drop_duplicates()
    dict_file_path = {}
    all_files = data["File Name"].tolist()
    dict_all = {}
    dict_index_number = {}
    each_row = []
    for file in all_files:
        dict_all[file] = data[data['File Name'] == file]['Keyword'].tolist()
        dict_index_number[file] = data[data['File Name'] == file]['Index Number'].tolist()[0]
        dict_file_path[file] = data[data['File Name'] == file]['File Path'].tolist()[0]
    workbook1 = xlrd.open_workbook(drl_path)
    sh = workbook1.sheet_by_index(0)
    for rownum in range(sh.nrows):
        rows = sh.row_values(rownum)
        if rows[5].strip().lower() == "it":
            sentence = rows[6]
            drl_number = int(rows[1])
            satisfied_output_file_name = {}
            satisfied_output_file_path = {}
            satisfied_output_index = {}
            flag = "No"
            relevant = 0
            # keys = file name
            for keys, values in dict_all.items():
                count = 0
                for m in values:
                    r1 = re.compile('|'.join([r'(?<!\w)%s(?!\w)' % re.escape(m)]), flags=re.I)
                    matches = r1.findall(sentence)
                    if matches:
                        count = count + 1
                if count >= 5:
                    flag = "Yes"
                    relevant = relevant + 1
                    if drl_number not in satisfied_output_file_name:
                        satisfied_output_file_name[int(drl_number)] = keys + ', '
                        satisfied_output_file_path[int(drl_number)] = dict_file_path[keys] + ", "
                        satisfied_output_index[int(drl_number)] = str(dict_index_number[keys]) + ", "
                    else:
                        satisfied_output_file_name[int(drl_number)] += keys + ", "
                        satisfied_output_file_path[int(drl_number)] += dict_file_path[keys] + ", "
                        satisfied_output_index[int(drl_number)] += str(dict_index_number[keys]) + ", "
            if (satisfied_output_file_name and satisfied_output_file_path \
                    and satisfied_output_index):
                dict1 = {}
                dict1["DRL #"] = drl_number
                dict1["Request"] = sentence
                dict1["Relevant Files Found?"] = flag
                dict1["Number of relevant files"] = relevant
                dict1["File Name(s)"] = satisfied_output_file_name[drl_number]
                dict1["Index Number(s)"] = satisfied_output_index[drl_number]
                dict1["File Path(s)"] = satisfied_output_file_path[drl_number]
                each_row.append(dict1)
            else:
                dict1 = {}
                dict1["DRL #"] = drl_number
                dict1["Request"] = sentence
                dict1["Relevant Files Found?"] = flag
                dict1["Number of relevant files"] = relevant
                dict1["Index Number(s)"] = ""
                dict1["File Path(s)"] = ""
                dict1["File Name(s)"] = ""
                each_row.append(dict1)

    dataframe = pd.DataFrame(each_row)
    dataframe = dataframe[
            ["DRL #", "Request", "Relevant Files Found?", "Number of relevant files", "File Name(s)", "Index Number(s)",
             "File Path(s)"]]
    # writing to "File Level Detail" sheet
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        dataframe.to_excel(writer, "DRL Details", index=False)
        writer.save()

    wb1 = openpyxl.load_workbook(output_file)
    ws = wb1['DRL Details']
    fillBack = PatternFill(start_color="00FFFF00", fill_type="solid")

    ws.column_dimensions['A'].width = 24.94
    ws.column_dimensions['B'].width = 24.94
    ws.column_dimensions['C'].width = 24.94
    ws.column_dimensions['D'].width = 24.94
    ws.column_dimensions['E'].width = 24.94
    ws.column_dimensions['F'].width = 24.94
    ws.column_dimensions['G'].width = 24.94
    ws.row_dimensions[1].height = 28.8

    for cell in ws["1:1"]:
        cell.fill = fillBack
    wb1.save(output_file)

def keyword_sql():
    try:
        db_connection_str = 'mysql+pymysql://root:admin@localhost/virtual_data_room'
        db_connection = create_engine(db_connection_str)
        df = pd.read_sql('SELECT keyword, category FROM additional_data_dictionary', con=db_connection)
        return df.drop_duplicates()
    except:
        print("mysql connection issue - ignoring the user input")
        sys.exit(1)


def cross_join(left, right):
    return left.assign(key=1).merge(right.assign(key=1), on='key').drop('key', 1)


def multi_processing(i):
    global keyword_dic
    print("file name ", i, keyword_dic)
    if i.endswith(".csv"):
        csv_handler(i)
    elif i.endswith((".xlsx", ".xlsm", ".xls", ".XLS")):
        print(datetime.datetime.now())
        xlsx_handler(i)
    elif i.endswith((".xlsb")):
        xlsb_handler(i)
    elif i.endswith(".msg"):
        msg_handler(i)
    elif i.endswith(".txt"):
        txt_handler(i)
    elif i.endswith(".rtf"):
        rtf_handler(i)
    elif i.endswith((".docx", ".doc")):
        print(datetime.datetime.now())
        doc_handler(i)
    elif i.endswith(".odt"):
        odt_handler(i)
    elif i.endswith((".pptx", "ppt")):
        print(datetime.datetime.now())
        ppt_handler(i)
    elif i.endswith((".PNG", ".png", ".JPEG", ".jpeg", ".JPG", ".jpg", ".gif", ".pnm", ".PNM")):
        img_handler(i)
    elif i.endswith((".pdf", ".PDF")):
        print(datetime.datetime.now())
        pdf_handler(i)
    elif i.endswith(".vsdx"):
        visio_handler(i)
    elif i.endswith((".tiff", ".tif", ".jfif", ".bmp", ".vsdx")):
        tiff_handler(i)
    else:
        list_details = folderName(i)
        outputWriterRawDetail(list_details[0], list_details[1], list_details[2], i, "Unknown", "", "",
                              "", "")

if __name__ == '__main__':
    try:
        file_count = 1
        file = open(intermediate_file)
        content = file.read()
        input_dict = ast.literal_eval(content)

        root = input_dict['VDR_Location']
        drl_file_path = input_dict['DRL_Location']


        fileList = []
        for path, subdirs, files in os.walk(root):
            for name in files:
                a = os.path.join(path, name)
                fileList.append(a)


        p = multiprocessing.Pool(3)
        p.map(multi_processing, fileList)


        """
        for file in fileList:
            multi_processing(file)
        """
        print("done")
        workbook.close()
        generatingFileLevelDetail(output_file)
        generatingKeywordDetails(output_file)
        generatingDRLDetails(drl_file_path, output_file)

    except KeyError as e:
        print("Either data is not present or column is missing", e)
        logger.critical("Either data is not present or column is missing", e)
    except PermissionError as e:
        print("Some files are open or not have permission to open - thus not accessible", e)
    except Exception as e:
        logger.info("Fatal error has occured - ", e)
    finally:
        print(memory_error_files)
        print(exception_files)
        logger.info("The end")
        logger.info("***********************************NEXT RUN************************************")
        logger.info("*******************************************************************************")
        logger.info("*******************************************************************************")
