# -*- coding: utf-8 -*-
"""
Created on Mon Jan 20 12:36:08 2020
@author: sbiswas149
"""

import csv
import xlrd
import re
import os
import xlsxwriter
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from pyxlsb import open_workbook
import numpy as np
from independentsoft.msg import Message
from striprtf.striprtf import rtf_to_text
import comtypes.client
import sys
import PyPDF2
from odf import text, teletype
from odf.opendocument import load
from sqlalchemy import create_engine
import pymysql
import ast
import configparser

outputCol = 0
outputRow = 1

try:
    config = configparser.ConfigParser()
    config.read('properties.ini')

    output_file = config['Directories']['output_file']
    temp_storage_path = config['Directories']['temp_storage']
    keyword_file = config['Directories']['keyword_file']
    intermediate_file = config['Directories']['intermediate_file']
except KeyError as e:
    print("Config file is empty")
    sys.exit(1)
except Exception as e:
    print("Check config file", e)
    sys.exit(1)

workbook = xlsxwriter.Workbook(output_file)
#sheet - Raw Details
header_format = workbook.add_format({
    'bold': True,
    'text_wrap': True,
    'valign': 'top',
    'fg_color': '#FFFF00',
    'border': 1})

worksheet1 = workbook.add_worksheet("Raw Details")
worksheet1.set_column(0, 8, 25)
worksheet1.write('A1','File Name', header_format)
worksheet1.write('B1','Index Number', header_format)
worksheet1.write('C1','Folder Name', header_format)
worksheet1.write('D1','File Path', header_format)
worksheet1.write('E1','File Type', header_format)
worksheet1.write('F1','Total Number of Page(s)/Slide(s)', header_format)
worksheet1.write('G1','Keyword', header_format)
worksheet1.write('H1','IT Category', header_format)
worksheet1.write('I1','Page/Slide Number', header_format)

def specialCharReplace(word):
    return re.sub('[^A-Za-z0-9\s]+', '', word)

def cell_header(a):
    if a <= 26:
        return chr(a+64)
    count = a//26 + 64
    rest = a%26 + 64
    return chr(count) + chr(rest)

def csv_handler(path):
    with open(path) as csvfile:
        readCSV = csv.reader(csvfile)
        count = 0
        page_count = 1
        for row in readCSV:
            # count is row number 
            count = count + 1
            for i in range(len(row)):
                #print(row[i] + " " + str(count) + " " + str(i))
                # row is representing each row  of the csv file - a list 
                list_details = folderName(path)
                cell_details = cell_header(i+1)+str(count)
                sheet_handler(list_details, cell_details, path, page_count, row[i])
    
def sheet_handler(list_details, cell_details, path, page_count, row):
    global r3, r4
    if isinstance(row, str):
        mm1 = [i for i in r1.findall(row) if i and isinstance(i, str)]
        if mm1:
            for keyword in mm1:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "IT Spend", cell_details)
        mm2 = [i for i in r2.findall(row) if i and isinstance(i, str)]                   
        if mm2:
            for keyword in mm2:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "IT Organization / Roles", cell_details)
                
        mm3 = [i for i in r3.findall(row) if i and isinstance(i, str)]
        #print(mm3)                 
        if mm3:
            for keyword in mm3:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "IT Applications", cell_details)
                             
        mm4 = [i for i in r4.findall(row) if i and isinstance(i, str)]
        if mm4:
            for keyword in mm4:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "IT Infrastructure", cell_details)
                        
        mm5 = [i for i in r5.findall(row) if i and isinstance(i, str)]
        if mm5:
            for keyword in mm5:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "IT Security and Controls", cell_details)
                        
        mm6 = [i for i in r6.findall(row) if i and isinstance(i, str)]
        if mm6:
            for keyword in mm6:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "IT Projects", cell_details)
                        
        mm7 = [i for i in r7.findall(row) if i and isinstance(i, str)]
        if mm7:
            for keyword in mm7:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "Target Technology Fitness Scorecard", cell_details)
                        
        mm8_1 = [i for i in r8.findall(row) if i and isinstance(i, str)]
        
        if mm8_1:
            for keyword in mm8_1:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "Misecellaneous", cell_details) 
        
        """        
        mm8_2 = [i for i in r8_2.findall(row) if i and isinstance(i, str)]
        print(mm8_2)
        if mm8_2:
            for keyword in mm8_2:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "Misecellaneous", cell_details)
        
        mm9 = [i for i in r9.findall(row) if i and isinstance(i, str)]                   
        if mm9:
            for keyword in mm9:
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], path, list_details[3], page_count,
                        keyword, "User Input", cell_details)
        """
    
  
def xlsx_handler(path):
    sheets_dict = pd.read_excel(path, sheet_name=None, header = None)
    sheet_count = len(sheets_dict)
    for name, sheet in sheets_dict.items():
        for index, row in sheet.iterrows():
            for i in range(len(row)):
                list_details = folderName(path)
                cell = cell_header(i+1)+str(index+1)
                cell_details = name + ": " + cell
                sheet_handler(list_details, cell_details, path, sheet_count, row[i])
                
def xlsb_handler(path): 
    sheet_count = len(open_workbook(path).sheets)
    with open_workbook(path) as wb:
        for sheetname in wb.sheets:
            row_number = 0
            with wb.get_sheet(sheetname) as sheet:
                for row in sheet.rows():
                    row_number = row_number + 1
                    for i in range(len(row)):
                        if row[i].v:
                            list_details = folderName(path)
                            cell = cell_header(i+1)+str(row_number)
                            cell_details = sheetname + ": " + cell
                            sheet_handler(list_details, cell_details, path, sheet_count, row[i].v) 
                            
def odt_handler(path):
    textdoc = load(path)
    allparas = textdoc.getElementsByType(text.P)
    list_details = folderName(path)
    for i in allparas:
        sheet_handler(list_details, 1, path, 1, str(i))
 
def pdf_handler(path, path_doc):
    list_details = folderName(path_doc)
    with open(path,'rb') as pdf_file:
        read_pdf = PyPDF2.PdfFileReader(pdf_file)
        number_of_pages = read_pdf.getNumPages()
        for page_number in range(number_of_pages):   # use xrange in Py2
            page = read_pdf.getPage(page_number)
            page_content = page.extractText()
            sheet_handler(list_details, page_number, path_doc, number_of_pages, page_content)

def doc_handler(path):
    global temp_storage_path
    print(temp_storage_path)
    wdFormatPDF = 17
    in_file = path
    out_file = os.path.abspath(temp_storage_path)
    word = comtypes.client.CreateObject('Word.Application')
    doc = word.Documents.Open(in_file)
    doc.SaveAs(out_file, FileFormat=wdFormatPDF)
    doc.Close()
    word.Quit()
    pdf_handler(temp_storage_path, path)
    os.remove(temp_storage_path)
                          
def msg_handler(path):
    content = str(Message(path).body)
    sub = str(Message(path).subject)
    list_details = folderName(path)
    sheet_handler(list_details, 1, path, 1, content)
    sheet_handler(list_details, 1, path, 1, sub)
    
def txt_handler(path):
    list_details = folderName(path)
    file = open(path,mode='r')
    content = file.read()
    sheet_handler(list_details, 1, path, 1, content)
    file.close()       
    
def rtf_handler(path):
    list_details = folderName(path)
    file = open(path,mode='r')
    rtf = file.read()
    content = rtf_to_text(rtf)
    sheet_handler(list_details, 1, path, 1, content)
    file.close()   
    
def index_number_verification(index):
    if "_" in index:
        index = index.split("_")[0]
    index = re.sub('[^0-9\.\s]+', '', index)
    if len(index) == 1:
        index = re.sub('[.]', " ", index)
    return index 
   
def folderName(path):
    #C:\Users\sb512911\Desktop\All\Screenshots\1.2.1 ILS Structure Chart.csv
    file_name = path.split("\\")[-1]
    index = file_name.split(" ")[0]   
    index_number = index_number_verification(index)
    folder_name = path.split(file_name)[0]
    file_type = path.split(".")[-1]   
    return [file_name, index_number, folder_name, file_type]

def outputWriterRawDetail(file_name, index_number, folder_name, file_path, file_type, pages, keyword, category, cell):
    global outputRow
    worksheet1.write(outputRow,0,file_name)
    worksheet1.write(outputRow,1,index_number)
    worksheet1.write(outputRow,2,folder_name)
    worksheet1.write(outputRow,3,file_path)
    worksheet1.write(outputRow,4,file_type)
    worksheet1.write(outputRow,5,pages)
    worksheet1.write(outputRow,6,keyword)
    worksheet1.write(outputRow,7,category)
    worksheet1.write(outputRow,8,cell)
    outputRow = outputRow + 1

# pass the additional parameter for relevance test - dynamically
def generatingFileLevelDetail(file_path):
    try:
        data = pd.read_excel(open(file_path,'rb'), sheet_name=0)
        if not data.empty:
            temp_data = data[["File Name","Index Number", "Folder Name", "File Type"]]
            temp_data["Scan Status"] = "Successful"
            file_level_detail = temp_data[temp_data['File Type'] != "Unknown"]\
                .drop_duplicates().sort_values('File Name')

            file_level_detail["Number of Keywords found"] = \
                    data[data["File Type"] != "Unknown"][['File Name']].sort_values('File Name').groupby(['File Name']).size().reset_index()[0].values

            file_level_detail["Number of Keyword Categories"] = \
                    data[data["File Type"] != "Unknown"][['File Name', 'IT Category']].drop_duplicates().groupby(['File Name']).size()\
                    .reset_index()[0].values

            intermediate1 = data[data["File Type"] != "Unknown"][['File Name','Keyword']].groupby(['File Name', 'Keyword']).size().reset_index()
            intermediate1[[0]] = intermediate1[[0]].astype(str).values
            intermediate1['joined'] = intermediate1[['Keyword', 0]].apply(lambda x: '('.join(x)+"), ", axis =1).values
            file_level_detail["Keyword Details"] = intermediate1[["File Name", "joined"]].groupby(["File Name"]).sum().values

            file_level_detail["Category Details"] = \
                data[data["File Type"] != "Unknown"].groupby("File Name").apply(lambda x : x["IT Category"].drop_duplicates().str.cat(sep=", ")).values

            file_level_detail["File Path"] = data[data["File Type"]!="Unknown"]["File Path"].drop_duplicates().values

            # Relevance check -- code here - assuming relevant count = 5
            mapping = {True: "Yes", False: "No"}
            file_level_detail["Relevant?"] = data[data["File Type"] != "Unknown"][["File Name", "Keyword"]].drop_duplicates()[["File Name"]].groupby(["File Name"]).size().\
                reset_index()[0].apply(lambda x: x>5).map(mapping).values

            temp_data_unknown = temp_data[["File Name","Index Number", "Folder Name", "File Type"]][temp_data['File Type'] == "Unknown"]
            temp_dict = {"Number of Keywords found" : 0, "Number of Keyword Categories":0, "Keyword Details" : "",\
                    "Category Details":"", "Relevant?":"", "File Path":"", "Scan Status": "Unsuccessful"}
            temp_df = pd.DataFrame([temp_dict])
            unknown_files = cross_join(temp_data_unknown, temp_df)
            file_level_detail = file_level_detail.append(unknown_files, ignore_index=True)
            writer = pd.ExcelWriter(output_file)
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                file_level_detail.to_excel(writer, "File Level Detail", index = False)
            writer.save()

            wb1 = openpyxl.load_workbook(output_file)
            ws = wb1['File Level Detail']
            fillBack = PatternFill(start_color="00FFFF00", fill_type="solid")

            ws.column_dimensions['A'].width = 24.94
            ws.column_dimensions['B'].width = 24.94
            ws.column_dimensions['C'].width = 24.94
            ws.column_dimensions['D'].width = 24.94
            ws.column_dimensions['E'].width = 24.94
            ws.column_dimensions['F'].width = 24.94
            ws.column_dimensions['G'].width = 24.94
            ws.column_dimensions['H'].width = 24.94
            ws.column_dimensions['I'].width = 24.94
            ws.column_dimensions['J'].width = 24.94
            ws.column_dimensions['K'].width = 24.94
            ws.row_dimensions[1].height  = 28.8

            for cell in ws["1:1"]:
                cell.fill = fillBack
            wb1.save(output_file)

    except:
        print("error occured while generating file level detail")
        sys.exit(1)

def generatingKeywordDetails(file_path):
    data = pd.read_excel(open(file_path,'rb'), sheet_name="Raw Details")
    try:
        if not data.empty:
        #distinct keywords
            keyword_details = data[["Keyword"]].drop_duplicates()

            # distinct keywords with the IT category
            keyword_details = data[["Keyword", "IT Category"]]\
                .drop_duplicates().merge(keyword_details, on="Keyword")

            #  Number of files with keywords
            keyword_details = data[["Keyword", "File Name"]].drop_duplicates().groupby(["Keyword"])\
                .size().reset_index().merge(keyword_details, on = "Keyword")

            # renaming
            keyword_details = keyword_details.\
                rename(columns={'IT Category': 'Keyword Category', 0: 'Number of files with keywords'})

            #Total keyword hits
            keyword_details = data[["File Name","Keyword"]].groupby(["Keyword"])\
                .size().reset_index().merge(keyword_details, on = "Keyword").rename(columns={0:"Total keyword hits"})

            #File(s) with maximum keyword hits and File Path(s)
            intermediate2 = data[["Keyword", "File Name", "File Path"]].groupby(["Keyword", "File Name", "File Path"])\
                .size().reset_index().merge(keyword_details, on = "Keyword").rename({0:"Count"}, axis = 1)\
                    [["File Name", "Keyword", "File Path", "Count"]]

            intermediate3 = intermediate2.groupby('Keyword')['Count'].apply(lambda x : x.eq(x.max()))

            intermediate4 = intermediate2.loc[intermediate3].groupby(['Keyword'])['File Name'].agg(', '.join).reset_index()

            intermediate5 = intermediate2.loc[intermediate3].groupby(['Keyword'])['File Path'].agg(', '.join).reset_index()

            keyword_details = keyword_details.merge(intermediate4, on = "Keyword")

            keyword_details = keyword_details.merge(intermediate5, on = "Keyword")

            keyword_details = keyword_details.rename(columns = {"File Path" : "File Path(s) with maximum keyword",\
                                    "File(s) with maximum keyword hits": "File Path(s) with maximum keyword"})

            # All file(s) with keyword hits
            keyword_details["File Path(s)"] = \
                data[["Keyword", "File Path"]].groupby(["Keyword"])\
                .apply(lambda x : x["File Path"].drop_duplicates().str.cat(sep=", ")).reset_index().\
                merge(keyword_details,on = "Keyword")[[0]].values

            # Index Number(s)
            keyword_details["Index Number(s)"] = \
                data[["Keyword", "Index Number"]].groupby("Keyword")\
                .apply(lambda x : x["Index Number"].drop_duplicates().str.cat(sep=", ")).reset_index()\
                .merge(keyword_details, on = "Keyword")[[0]].values

            # All the file names which has the keyword
            keyword_details["File Name(s)"] = \
                data[["Keyword", "File Name"]].groupby("Keyword")\
                .apply(lambda x : x["File Name"].drop_duplicates().str.cat(sep=", ")).reset_index()\
                .merge(keyword_details, on = "Keyword")[[0]].values

            keyword_details = keyword_details.rename(columns = {"File Name" : "File Name(s) with maximum keyword",\
                                            "File Name(s)": "All file(s) with keyword hits"})

            keyword_details = keyword_details[["Keyword", "Keyword Category", "Number of files with keywords",\
                "Total keyword hits", "File Name(s) with maximum keyword", "File Path(s) with maximum keyword",\
                "All file(s) with keyword hits", "Index Number(s)", "File Path(s)"]]

            # writing to "File Level Detail" sheet
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                keyword_details.to_excel(writer, "Keyword Details", index = False)
            writer.save()

            wb1 = openpyxl.load_workbook(output_file)
            ws = wb1['Keyword Details']
            fillBack = PatternFill(start_color="00FFFF00", fill_type = "solid")

            ws.column_dimensions['A'].width = 24.94
            ws.column_dimensions['B'].width = 24.94
            ws.column_dimensions['C'].width = 24.94
            ws.column_dimensions['D'].width = 24.94
            ws.column_dimensions['E'].width = 24.94
            ws.column_dimensions['F'].width = 24.94
            ws.column_dimensions['G'].width = 24.94
            ws.column_dimensions['H'].width = 24.94
            ws.column_dimensions['I'].width = 24.94
            ws.row_dimensions[1].height = 28.8

            for cell in ws["1:1"]:
                cell.fill = fillBack
            wb1.save(output_file)

    except Exception as e:
        print("Error occured while generating DRL details", e)
        
def generatingDRLDetails(drl_path, output_path):
    drl_content = pd.read_excel(open(output_path,'rb'), None)
    if 'File Level Detail' in drl_content.keys():
        df = pd.read_excel(output_path, sheet_name = "File Level Detail")
        data = df[df["Scan Status"] != "Unsuccessful"][["File Name", "Index Number", "File Path", "Keyword Details"]]
        dict_val = {}
        dict_file_path = {}
        dict_index_number = {}
        each_row = []
        for index, row in data.iterrows():
            dict_index_number[row[0]] = row[1]
            dict_file_path[row[0]] = row[2]
            dict_val[row[0]] = row[3].split(",")
        
        workbook1 = xlrd.open_workbook(drl_path)
        sh = workbook1.sheet_by_index(0)
        for rownum in range(sh.nrows):
            rows = sh.row_values(rownum)
            if rows[5].strip().lower() == "it":
                sentence = rows[6]
                drl_number = int(rows[1])
                satisfied_output_file_name = {}
                satisfied_output_file_path = {}
                satisfied_output_index = {}
                flag = "No"
                relevant = 0
                for keys,values in dict_val.items():
                    count = 0
                    for m in values:
                        if m[:-3].lower() in sentence.lower():
                            if m.strip():
                                count = count + int(m.strip()[-2:-1])
                    if count >= 5:
                        flag = "Yes"
                        relevant = relevant + 1
                        if drl_number not in satisfied_output_file_name:
                            satisfied_output_file_name[int(drl_number)] =  keys + ', '
                            satisfied_output_file_path[int(drl_number)] = dict_file_path[keys] + ", "
                            satisfied_output_index[int(drl_number)] = str(dict_index_number[keys]) + ", "
                        else:
                            satisfied_output_file_name[int(drl_number)] += keys + ", "
                            satisfied_output_file_path[int(drl_number)] += dict_file_path[keys] + ", "
                            satisfied_output_index[int(drl_number)] += str(dict_index_number[keys]) + ", "
                if (satisfied_output_file_name and satisfied_output_file_path\
                      and satisfied_output_index):
                    #writingDRL(drl_number, sentence, satisfied_output_file_name[drl_number], satisfied_output_file_path[drl_number],\
                        #satisfied_output_index[drl_number], flag, relevant) 
                    dict1 = {}
                    dict1["DRL #"] = drl_number
                    dict1["Request"] = sentence
                    dict1["Relevant Files Found?"] = flag
                    dict1["Number of relevant files"] = relevant
                    dict1["File Name(s)"] = satisfied_output_file_name[drl_number]
                    dict1["Index Number(s)"] = satisfied_output_index[drl_number]
                    dict1["File Path(s)"] = satisfied_output_file_path[drl_number]
                    each_row.append(dict1)
                else:
                    dict1 = {}
                    dict1["DRL #"] = drl_number
                    dict1["Request"] = sentence
                    dict1["Relevant Files Found?"] = flag
                    dict1["Number of relevant files"] = relevant
                    dict1["Index Number(s)"] = ""
                    dict1["File Path(s)"] = ""
                    dict1["File Name(s)"] = ""
                    each_row.append(dict1)
                              
        dataframe = pd.DataFrame(each_row) 
        dataframe = dataframe[["DRL #", "Request", "Relevant Files Found?", "Number of relevant files", "File Name(s)", "Index Number(s)", "File Path(s)"]]
          # writing to "File Level Detail" sheet

        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            dataframe.to_excel(writer, "DRL Details", index = False)    
            writer.save()

        wb1 = openpyxl.load_workbook(output_file)
        ws = wb1['DRL Details']
        fillBack = PatternFill(start_color="00FFFF00", fill_type="solid")

        ws.column_dimensions['A'].width = 24.94
        ws.column_dimensions['B'].width = 24.94
        ws.column_dimensions['C'].width = 24.94
        ws.column_dimensions['D'].width = 24.94
        ws.column_dimensions['E'].width = 24.94
        ws.column_dimensions['F'].width = 24.94
        ws.column_dimensions['G'].width = 24.94
        ws.row_dimensions[1].height = 28.8

        for cell in ws["1:1"]:
            cell.fill = fillBack
        wb1.save(output_file)
            
def keyword_sql():
    try:
        db_connection_str = 'mysql+pymysql://root:admin@localhost/virtual_data_room'
        db_connection = create_engine(db_connection_str)
        df = pd.read_sql('SELECT keyword, category FROM additional_data_dictionary', con=db_connection)
        return df.drop_duplicates()
    except:
        print("mysql connection issue - ignoring the user input")
        sys.exit(1)

def cross_join(left, right):
    return left.assign(key=1).merge(right.assign(key=1), on='key').drop('key', 1)

# triggering the application
if __name__ == '__main__':
    try:
        #df = keyword_sql()
        wb = xlrd.open_workbook(keyword_file)
        sh = wb.sheet_by_index(0)
        file = open(intermediate_file)
        content = file.read()
        input_dict = ast.literal_eval(content)   
        
        root = input_dict['VDR_Location']
        drl_file_path = input_dict['DRL_Location']
       
        it_spend = [i for i in sh.col_values(1, start_rowx = 1) if i]
        it_org = [i for i in sh.col_values(2, start_rowx = 1) if i]
        it_apps = [i for i in sh.col_values(3, start_rowx = 1, end_rowx = 40) if i]
        it_infra = [i for i in sh.col_values(4, start_rowx = 1, end_rowx = 40) if i]
        it_security = [i for i in sh.col_values(5, start_rowx = 1) if i]
        it_projects = [i for i in sh.col_values(6, start_rowx = 1) if i]
        target_tech_fitness = [i for i in sh.col_values(7, start_rowx = 1) if i]
        misecellaneous = [i for i in sh.col_values(8, start_rowx = 1) if i]
        #misecellaneous2 = [i for i in sh.col_values(9, start_rowx = 1) if i]
        #user_input = df[df["category"] == "User_Input"]["keyword"].values.tolist()
       
        r1 = re.compile('|'.join([r'\b%s\b' % w for w in it_spend]), flags=re.I)
        r2 = re.compile('|'.join([r'\b%s\b' % w for w in it_org]), flags=re.I)
        r3 = re.compile('|'.join([r'\b%s\b' % w for w in it_apps]), flags=re.I)
        #print(r3)
        r4 = re.compile('|'.join([r'\b%s\b' % w for w in it_infra]), flags=re.I)
        r5 = re.compile('|'.join([r'\b%s\b' % w for w in it_security]), flags=re.I)
        r6 = re.compile('|'.join([r'\b%s\b' % w for w in it_projects]), flags=re.I)
        r7 = re.compile('|'.join([r'\b%s\b' % w for w in target_tech_fitness]), flags=re.I)
        r8 = re.compile('|'.join([r'\b%s\b' % w for w in misecellaneous]), flags=re.I)
        #print(r8_1)
        #r8_2 = re.compile('|'.join([r'\b%s\b' % w for w in misecellaneous2]), flags=re.I)
        #print(r8_2)
        #r9 = re.compile('|'.join([r'\b%s\b' % w for w in user_input]), flags=re.I)
       
        fileList = []
        for path, subdirs, files in os.walk(root):
            for name in files:
                a = os.path.join(path, name)
                fileList.append(a)
               
        for i in fileList:
            if i.endswith(".csv"):
                csv_handler(i)   
            elif i.endswith((".xlsx", ".xlsm", ".xls", ".XLS")):
                xlsx_handler(i)
            elif i.endswith((".xlsb")):
                xlsb_handler(i)
            elif i.endswith(".msg"):
                msg_handler(i)
            elif i.endswith(".txt"):
                txt_handler(i)
            elif i.endswith(".rtf"):
                rtf_handler(i)
            elif i.endswith((".docx", ".doc")):
                doc_handler(i)
            elif i.endswith(".odt"):
                odt_handler(i)
            else:
                list_details = folderName(i)
                outputWriterRawDetail(list_details[0], list_details[1], list_details[2], i, "Unknown", "", "",
                                      "", "")

        workbook.close()
        generatingFileLevelDetail(output_file)
        generatingKeywordDetails(output_file)
        generatingDRLDetails(drl_file_path, output_file)
        
        print("The end")
    except KeyError as e:
        print("Either data is not present or column is missing")
    except PermissionError as e:
        print("Some files are open - thus not accessible")
